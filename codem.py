import PyPDF2
from tkinter import filedialog, Tk
from openpyxl import Workbook
from openpyxl.styles import Alignment
import re

def extract_empenho_and_historic_from_pdf(pdf_path):
    pdf_file = open(pdf_path, 'rb')
    pdf_reader = PyPDF2.PdfReader(pdf_file)

    empenho_historics = []
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        text = page.extract_text()
        empenho = re.search(r'2024NE(\d{6})', text)
        if empenho:
            empenho = empenho.group(1)
            # Procurando pelo campo "Histórico" em cada página
            hist_index = text.find("Histórico")
            if hist_index != -1:
                # Se encontrado, pegar o texto a partir desse ponto até o fim da página
                historic = text[hist_index+len("Histórico"):].strip()
                # Remover a parte do texto que vai desde o início do termo "Programa Trabalho" até o fim do documento
                prog_trabalho_index = historic.find("Programa Trabalho")
                if prog_trabalho_index != -1:
                    historic = historic[:prog_trabalho_index].strip()
                empenho_historics.append((empenho, historic))
    
    pdf_file.close()
    return empenho_historics

def save_to_excel(empenho_historics, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historics"

    # Definindo os cabeçalhos das colunas
    ws['A1'] = "Número do Empenho"
    ws['B1'] = "Histórico"

    for idx, (empenho, historic) in enumerate(empenho_historics, start=2):
        ws[f'A{idx}'] = empenho
        ws[f'B{idx}'] = historic
        ws[f'A{idx}'].alignment = Alignment(wrapText=True)
        ws[f'B{idx}'].alignment = Alignment(wrapText=True)

    wb.save(excel_path)

def open_file_dialog():
    root = Tk()
    root.withdraw()  # Esconde a janela principal

    # Estilizando a janela de seleção de arquivos
    root.tk_setPalette(background='black', foreground='green4')

    file_path = filedialog.askopenfilename(title="Selecione o arquivo PDF",
                                            filetypes=[("PDF files", "*.pdf")])
    return file_path

if __name__ == "__main__":
    empenho_historics = []

    while True:
        print("Selecione o próximo arquivo PDF:")
        pdf_path = open_file_dialog()

        if not pdf_path:
            print("Nenhum arquivo selecionado. Encerrando o programa.")
            break

        print("Extraindo números de empenho e históricos...")
        empenho_historics.extend(extract_empenho_and_historic_from_pdf(pdf_path))

        print("Números de empenho e históricos extraídos com sucesso!")

    if empenho_historics:
        print("Selecione onde deseja salvar o arquivo Excel:")
        excel_path = filedialog.asksaveasfilename(title="Salvar históricos como",
                                                    defaultextension=".xlsx",
                                                    filetypes=[("Excel files", "*.xlsx")])

        if excel_path:
            save_to_excel(empenho_historics, excel_path)
            print(f"Os números de empenho e históricos foram salvos com sucesso no arquivo '{excel_path}'.")
        else:
            print("Nenhum local de salvamento selecionado. Os números de empenho e históricos não foram salvos.")
    else:
        print("Nenhum número de empenho e histórico encontrado nos arquivos PDF. Encerrando o programa.")
